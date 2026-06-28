import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO

def generate_maintenance_plan_excel(plan_data: dict) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan de Maintenance"
    
    header_fill = PatternFill(start_color="163323", end_color="163323", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Horizon", "ID Action", "Ouvrage", "Nom", "Coût", "Gain Risque"]
    ws.append(headers)
    
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        
    for horizon, actions in plan_data.items():
        for a in actions:
            # action is a dict or ActionResponse model
            if not isinstance(a, dict):
                a = a.model_dump()
            ws.append([
                horizon,
                a.get('id'),
                a.get('ouvrage_id'),
                a.get('nom'),
                a.get('cout_estime'),
                a.get('gain_risque')
            ])
            
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
